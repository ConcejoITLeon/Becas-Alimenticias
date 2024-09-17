from flask import flash,Flask,render_template,request,redirect,url_for,jsonify,session,send_file
from flask_mysqldb import MySQL
from datetime import datetime
from datetime import timedelta
from app import create_app
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd
import io
import calendar

app,conexion=create_app()

@app.route('/')
def index():

    return render_template('inicio_sesion.html')

@app.route('/login', methods=['POST'])
def inicio_sesion():
    if request.method == 'POST':
        
        fecha_actual=datetime.now().strftime('%A %d %B')
        fecha_actual=fecha_actual.title()
        fecha_db_actual=datetime.now().strftime('%Y-%m-%d')
        numero_control = request.form['numcontrol']
        contrasena = request.form['pass']

        
        cursor=conexion.connection.cursor()
        cursor.execute("SELECT * FROM becarios WHERE numcontrol = %s", (numero_control,))
        usuario = cursor.fetchone()
        
        if usuario:
            cursor.execute("SELECT * FROM becarios WHERE numcontrol=%s AND pass=%s",(numero_control,contrasena,))
            
            us_pas=cursor.fetchone()
            
            
            if us_pas:
                if us_pas[3]==contrasena:
                    if us_pas[6]:
                        session.permanent = True
                        session['user'] = numero_control  # Guarda el número de control en la sesión
                        session['nombre_usuario']=us_pas[1]
                        session['fecha']=fecha_db_actual
                        session['fecha_letras']=fecha_actual
                        session['rol'] = us_pas[5]
                        cursor.execute("SELECT numcontrol FROM cobro_beca WHERE numcontrol=%s AND fecha=%s", (numero_control, fecha_db_actual,))
                        cobro_hecho = cursor.fetchone()
                        cursor.execute("SELECT grupo FROM becarios WHERE numcontrol=%s", (numero_control,))
                        grupo = cursor.fetchone()
                        
                        if grupo:
                            grupo=grupo[0]
                            dias_cobro = set()
                            if grupo == 'A':
                                dias_cobro = {0, 1, 2, 3, 4}  # Lunes a Viernes
                            elif grupo == 'B':
                                dias_cobro = {0, 2, 4}  # Lunes, Miércoles, Viernes
                            elif grupo == 'C':
                                dias_cobro = {1, 3}  # Martes, Jueves

                            dia_semana = datetime.now().weekday()
                            puede_cobrar = dia_semana in dias_cobro
                            
                            if session['rol']=='admin':
                                return redirect(url_for('resumen.resumen'))
                            else:
                                if cobro_hecho:
                                    return render_template('main_cobro.html', nombre_usuario=us_pas[1],fecha_actual=fecha_actual, cobro_realizado=True, fecha_db_actual=fecha_db_actual,puede_cobrar=puede_cobrar)
                                elif not puede_cobrar:
                                     return render_template('main_cobro.html', nombre_usuario=us_pas[1], fecha_actual=fecha_actual, cobro_realizado=False, fecha_db_actual=fecha_db_actual, puede_cobrar=puede_cobrar)
                                else:
                                    return render_template('main_cobro.html', nombre_usuario=us_pas[1],fecha_actual=fecha_actual, cobro_realizado=False, fecha_db_actual=fecha_db_actual,puede_cobrar=puede_cobrar)
                        cursor.close()
                        
                        
                    else:
                        cursor.close()
                        return render_template('inicio_sesion.html', error= 'Tu cuenta no se ha activado')
            else:
                cursor.close()
                return render_template('inicio_sesion.html',error='Contraseña Incorrecta')
        else:
            cursor.close()
            return render_template('inicio_sesion.html',error='Número de Control Incorrecto')

@app.route('/logout')
def logout():
    session.pop('user',None)
    return redirect(url_for('index'))

@app.route('/alta_becario', methods=['GET', 'POST'])
def alta_becario():
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        numcontroles = request.form.getlist('numcontrol')

        try:
            cursor=conexion.connection.cursor()
            cursor.executemany("UPDATE becarios SET activo = TRUE WHERE numcontrol= %s", [(numcontrol,) for numcontrol in numcontroles])
            conexion.connection.commit()
            cursor.close()
            flash('Becario añadido correctamente', 'success')
            return redirect(url_for('alta_becario'))
        except Exception as e:
            return render_template('alta_becario.html', error=str(e))

    try:
        cursor=conexion.connection.cursor()
        cursor.execute("SELECT nombre,numcontrol,campus FROM becarios WHERE activo = FALSE ORDER BY campus,nombre")
        usuarios_pendientes = cursor.fetchall()
        cursor.close()

        return render_template('alta_becario.html', usuarios_pendientes=usuarios_pendientes)    
    except Exception as e:
        return render_template('alta_becario.html', error=str(e))

@app.route('/alta_admin', methods=['GET', 'POST'])
def alta_admin():
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        numcontroles = request.form.getlist('numcontrol')

        try:
            cursor=conexion.connection.cursor()
            cursor.executemany("UPDATE becarios SET rol = 'admin', activo = true WHERE numcontrol= %s", [(numcontrol,) for numcontrol in numcontroles])
            conexion.connection.commit()
            cursor.close()
            return redirect(url_for('alta_admin'))
        except Exception as e:
            return render_template('listar_alta_admin.html', error=str(e))

    try:
        cursor=conexion.connection.cursor()
        cursor.execute("SELECT nombre,numcontrol FROM becarios WHERE rol = 'user' ORDER BY nombre")
        usuarios_pendientes = cursor.fetchall()
        cursor.close()

        return render_template('listar_alta_admin.html', usuarios_pendientes=usuarios_pendientes)    
    except Exception as e:
        return render_template('listar_alta_admin.html', error=str(e))

@app.route('/baja_admin', methods=['GET', 'POST'])
def baja_admin():
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        numcontroles = request.form.getlist('numcontrol')

        try:
            cursor=conexion.connection.cursor()
            cursor.executemany("UPDATE becarios SET rol = 'user', activo = false WHERE numcontrol= %s", [(numcontrol,) for numcontrol in numcontroles])
            conexion.connection.commit()
            cursor.close()
            return redirect(url_for('baja_admin'))
        except Exception as e:
            return render_template('listar_baja_admin.html', error=str(e))

    try:
        cursor=conexion.connection.cursor()
        cursor.execute("SELECT nombre,numcontrol FROM becarios WHERE rol = 'admin' ORDER BY nombre")
        usuarios_pendientes = cursor.fetchall()
        cursor.close()

        return render_template('listar_baja_admin.html', usuarios_pendientes=usuarios_pendientes)    
    except Exception as e:
        return render_template('listar_baja_admin.html', error=str(e))
 
@app.route('/baja_usuario', methods=['GET', 'POST'])
def baja_usuario():
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        numcontroles = request.form.getlist('numcontrol')

        try:
            cursor=conexion.connection.cursor()
            cursor.executemany("DELETE FROM becarios WHERE numcontrol= %s", [(numcontrol,) for numcontrol in numcontroles])
            conexion.connection.commit()
            cursor.executemany("DELETE FROM cobro_beca WHERE numcontrol= %s", [(numcontrol,) for numcontrol in numcontroles])
            conexion.connection.commit()
            cursor.close()
            return redirect(url_for('baja_usuario'))
        except Exception as e:
            return render_template('listar_baja_usuario.html', error=str(e))

    try:
        cursor=conexion.connection.cursor()
        cursor.execute("SELECT nombre,numcontrol FROM becarios WHERE rol = 'user' ORDER BY nombre")
        usuarios_pendientes = cursor.fetchall()
        cursor.close()

        return render_template('listar_baja_usuario.html', usuarios_pendientes=usuarios_pendientes)    
    except Exception as e:
        return render_template('listar_baja_usuario.html', error=str(e))
        
@app.route('/usuarios')
def listar_usuarios():
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))

    cursor = conexion.connection.cursor()
    cursor.execute("SELECT id,nombre, numcontrol, campus FROM becarios ORDER BY campus,nombre")
    usuarios = cursor.fetchall()
    cursor.close()
    
    
    return render_template('listar_usuarios.html', usuarios=usuarios)

@app.route('/rechazar_usuario/<numcontrol>', methods=['GET','POST'])
def rechazar_usuario(numcontrol):
    try:
        
        cursor = conexion.connection.cursor()

        # Consulta para eliminar el usuario de la tabla
        query = "DELETE FROM becarios WHERE numcontrol = %s"
        cursor.execute(query, (numcontrol,))
        conexion.connection.commit()

        # Mensaje de éxito
        flash('Usuario rechazado y eliminado correctamente', 'success')

    except Exception as e:
        # Manejo de errores y rollback de la transacción en caso de error
        conexion.connection.rollback()
        flash(f'Error al rechazar el usuario: {str(e)}', 'danger')

    finally:
        cursor.close()

    # Redirigir de vuelta a la página de activación de usuarios
    return redirect(url_for('alta_becario'))

@app.route('/editar_usuario/<int:id>', methods=['GET','POST'])
def editar_usuario(id):
    if 'user' not in session or session.get('rol') != 'admin':
        return redirect(url_for('index'))
    
    
    cursor=conexion.connection.cursor()
    cursor.execute("SELECT nombre,numcontrol,campus,pass FROM becarios WHERE id = %s", (id,))
    usuario = cursor.fetchone()
    

    if request.method == 'POST':
        nueva_contrasena = request.form['contrasena']
        cursor.execute("UPDATE becarios SET pass = %s WHERE id = %s", (nueva_contrasena,id))
        conexion.connection.commit()
        cursor.close()
        return redirect(url_for('listar_usuarios'))
    
    cursor.close()
    return render_template('editar_usuario.html', usuario=usuario,id=id)

@app.route('/registro', methods=['POST'])
def registro_usuario():
    if request.method == 'POST':
        nombre_completo = request.form['nombre'].title()
        numero_control = request.form['numcontrol']
        constrasena = request.form['pass']
        campus = request.form['campus']
        grupo= request.form['grupo']

        try:
            cursor= conexion.connection.cursor()
            cursor.execute("SELECT numcontrol FROM becarios WHERE numcontrol = %s", (numero_control,))
            usuario_existente=cursor.fetchone()

            if usuario_existente:
                return render_template('inicio_sesion.html', error='El número de control ya existe. Porfavor, intenta con otro numero de control.')
            
            cursor.execute("INSERT INTO becarios (nombre, numcontrol, pass, campus,grupo) VALUES (%s, %s, %s, %s,%s)",(nombre_completo,numero_control,constrasena,campus,grupo))
            conexion.connection.commit()
            cursor.close()
            return redirect(url_for('confirmacion_registro', numero_control=numero_control))
        except Exception as e:
            return redirect(url_for('404')),404

@app.route('/cobro_beca', methods=['GET','POST'])
def cobro_beca():
    if 'user' not in session:
        return redirect(url_for('index'))
    
    usuario= session.get('user')
    nombre_usuario= session.get('nombre_usuario')
    fecha=session.get('fecha')
    fecha_letras=session.get('fecha_letras')
    
    if request.method == 'POST':
        
       # nombre_usuario = request.form['nombre_usuario']
        #fecha = request.form['fecha']
       
        try:
            cursor= conexion.connection.cursor()
            cursor.execute("SELECT numcontrol,grupo FROM becarios WHERE nombre = %s", (nombre_usuario,))
            consulta_ncon=cursor.fetchone()

            if consulta_ncon is None:
                cursor.close()
                return render_template('main_cobro.html',error='No se encontró el usuario con ese nombre')
            
            #Encuentra numero control de usuario
            num_control,grupo=consulta_ncon
            
            dias_cobro = set()
            if grupo == 'A':
                dias_cobro = {0, 1, 2, 3, 4}  # Lunes a Viernes
            elif grupo == 'B':
                dias_cobro = {0, 2, 4}  # Lunes, Miércoles, Viernes
            elif grupo == 'C':
                dias_cobro = {1, 3}  # Martes, Jueves
            
            dia_semana = datetime.strptime(fecha, '%Y-%m-%d').weekday()
            
            puede_cobrar=dia_semana in dias_cobro
            
            #Verifica si ya cobró en el día actual
            cursor= conexion.connection.cursor()
            cursor.execute("SELECT numcontrol FROM cobro_beca WHERE numcontrol=%s AND fecha=%s", (num_control,fecha,))
            cobro_hecho=cursor.fetchone()

            #Si hay cobro se bloquea botón cobro y muestra mensaje
            if cobro_hecho:
                cursor.close()

                return render_template('main_cobro.html', error='El cobro ya ha sido realizado',nombre_usuario=nombre_usuario, fecha_actual=fecha)
            
            if not puede_cobrar:
                cursor.close()
                return render_template('main_cobro.html', error='Hoy no es un día de cobro para tu grupo', nombre_usuario=nombre_usuario, fecha_actual=fecha, puede_cobrar=puede_cobrar)

            #Si no hay cobro registra en DB el cobro
            cursor.execute("INSERT INTO cobro_beca (numcontrol,fecha) VALUES (%s,%s)",(num_control,fecha,))
            conexion.connection.commit()
            cursor.close()
            

            return redirect(url_for('confirmacion_cobro', nombre_usuario=nombre_usuario,fecha=fecha_letras,num_control=num_control))
        
        except  Exception as e:
            return render_template('main_cobro.html',error=str(e))
    else:
        if usuario:
            fecha=datetime.now().strftime('%Y-%m-%d')
            try:
                cursor = conexion.connection.cursor()
                cursor.execute("SELECT numcontrol,grupo FROM becarios WHERE nombre = %s", (nombre_usuario,))
                consulta_ncon = cursor.fetchone()
                
                if consulta_ncon is None:
                    cursor.close()
                    return render_template('main_cobro.html',error='No se encontró el usuario con ese nombre')
                
                #Encuentra numero control de usuario
                num_control,grupo=consulta_ncon
                
                dias_cobro = set()
                if grupo == 'A':
                    dias_cobro = {0, 1, 2, 3, 4}  # Lunes a Viernes
                elif grupo == 'B':
                    dias_cobro = {0, 2, 4}  # Lunes, Miércoles, Viernes
                elif grupo == 'C':
                    dias_cobro = {1, 3}  # Martes, Jueves
                
                dia_semana = datetime.strptime(fecha, '%Y-%m-%d').weekday()
                
                puede_cobrar=dia_semana in dias_cobro
                
                #Verifica si ya cobró en el día actual
                cursor= conexion.connection.cursor()
                cursor.execute("SELECT numcontrol FROM cobro_beca WHERE numcontrol=%s AND fecha=%s", (num_control,fecha,))
                cobro_hecho=cursor.fetchone()

                #Si hay cobro se bloquea botón cobro y muestra mensaje
                if cobro_hecho:
                    cursor.close()

                    return render_template('main_cobro.html',nombre_usuario=nombre_usuario, fecha_actual=fecha, cobro_realizado=cobro_hecho)
                
                elif not puede_cobrar:
                    cursor.close()
                    return render_template('main_cobro.html', nombre_usuario=nombre_usuario, fecha_actual=fecha, puede_cobrar=puede_cobrar)


                else:
                    return render_template('main_cobro.html', nombre_usuario=nombre_usuario,fecha_actual=fecha, cobro_realizado=False, puede_cobrar=puede_cobrar)
            
        
            except  Exception as e:
                return render_template('main_cobro.html',error=str(e))
        else:
            return render_template('main_cobro.html', error='No hay información disponible para mostrar.')
        
    
@app.route('/confirmacion_cobro/<nombre_usuario>/<fecha>')
def confirmacion_cobro(nombre_usuario,fecha):
    return render_template('confirmacion_cobro.html',nombre=nombre_usuario,fecha=fecha)
    
@app.route('/confirmacion_registro/<numero_control>')
def confirmacion_registro(numero_control):
    cursor=conexion.connection.cursor()

    cursor.execute("SELECT nombre,numcontrol,pass,campus,grupo FROM becarios WHERE numcontrol = %s", (numero_control,))
    datos=cursor.fetchone()
    cursor.close()

    if datos:
        nombre_completo,numero_control,contra,campus,grupo = datos
        return render_template('registro_exitoso.html',nombre_completo=nombre_completo, numero_control=numero_control, contra=contra,campus=campus,grupo=grupo)

@app.route('/justificaciones_pendientes',methods=['GET','POST'])
def justificaciones_pendientes():
    cursor=conexion.connection.cursor()
    cursor.execute("SELECT * FROM justificaciones")
    justificaciones=cursor.fetchall()
    
    cursor.close()

    return render_template('justificar_admin.html', justificaciones=justificaciones)

@app.route('/aceptar_justificacion/<int:id>', methods=['POST','GET'])
def aceptar_justificacion(id):
    cursor=conexion.connection.cursor()
    cursor.execute("SELECT numcontrol FROM justificaciones WHERE id=%s",(id,))
    resultado = cursor.fetchone()

    if resultado:
        numcontrol = resultado[0]

        cursor.execute("""
            SELECT COUNT(*) FROM cobro_beca
            WHERE numcontrol = %s AND fecha = (SELECT fecha FROM justificaciones WHERE id = %s)
            """,(numcontrol,id))
        
        if cursor.fetchone()[0] == 0:
            cursor.execute("""
                INSERT INTO cobro_beca (numcontrol,fecha)
                SELECT numcontrol,fecha
                FROM justificaciones
                WHERE id = %s
            """,(id,))
            conexion.connection.commit()
        
        cursor.execute("DELETE FROM justificaciones WHERE id = %s",(id,))
        conexion.connection.commit()
        flash('Justificación aceptada y estado actualizado.', 'success')
    else:
        flash('Justificación no encontrada.', 'danger')

    cursor.close()
    return redirect(url_for('justificaciones_pendientes'))

@app.route('/rechazar_justificacion/<int:id>',methods=['POST','GET'])
def rechazar_justificacion(id):
    cursor=conexion.connection.cursor()
    
    cursor.execute("DELETE FROM justificaciones WHERE id = %s",(id,))
    conexion.connection.commit()

    flash('Justificación rechazada y eliminada.', 'success')
    cursor.close()
    return redirect(url_for('justificaciones_pendientes'))


def guardar_archivo_db(nombre_archivo, archivo):
    with app.app_context():
        cursor=conexion.connection.cursor()

        try:
            cursor.execute("""
                        INSERT INTO respaldo_calendario (nombre_archivo,archivo)
                        VALUES (%s, %s)
                        """,(nombre_archivo,archivo))
            conexion.connection.commit()
        except Exception as e:
            conexion.connection.rollback()
            print(f"Error al guardar el archivo en la base de datos: {e}")
        finally:
            cursor.close()

def iniciar_scheduler():
    with app.app_context():
        scheduler=BackgroundScheduler()
        #scheduler.add_job(func=crear_respaldo_semanal, trigger='interval', minutes=1)
        scheduler.add_job(func=crear_respaldo_semanal, trigger='cron', day_of_week='tue',hour=10,minute=00)
        scheduler.start()

def crear_respaldo_semanal():
    with app.app_context():
        resumen_data=obtener_resumen_data()
        
        ahora = datetime.now()
        primer_dia = ahora.replace(day=1)
        ultimo_dia = ahora.replace(day=calendar.monthrange(ahora.year, ahora.month)[1])
        
        dias_laborales = set()
        while primer_dia <= ultimo_dia:
            if primer_dia.weekday() < 5:
                dias_laborales.add(primer_dia.day)
            primer_dia += timedelta(days=1)

        rows = []

        max_faltas = {
            'A': 5,
            'B': 3,
            'C': 2
        }

        for campus, grupos in resumen_data.items():
            for grupo, becarios in sorted(grupos.items(), key=lambda x: x[0]):
                becarios_ordenados = sorted(becarios, key=lambda b: b['nombre'])
                for becario in becarios_ordenados:
                    fila = [campus, grupo, becario['nombre'], becario['numcontrol']]
                    faltas = 0

                    # Añadir el estado de cada día del mes a la fila
                    for dia in dias_laborales:
                        estado_dia = becario['dias_cobro'].get(dia, 'no-aplica')
                        if estado_dia == 'no-cobrado':
                            faltas += 1
                        fila.append(estado_dia)
                    fila.append(faltas)
                    rows.append(fila)

        # Crear nombres de columna dinámicamente para Excel
        column_names = ['Campus', 'Grupo', 'Becario', 'Número de Control'] + [dia for dia in dias_laborales] + ['Faltas']

        # Crear DataFrame con las filas y columnas correctas
        df = pd.DataFrame(rows, columns=column_names)
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, sheet_name='Resumen de Cobros', index=False)
        
        workbook = writer.book
        sheet = workbook['Resumen de Cobros']

        green_fill = PatternFill(start_color="00b900", end_color="00b900", fill_type="solid")
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        gray_fill= PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        red_fill= PatternFill(start_color="FF6F65", end_color="FF6F65", fill_type="solid")
        wine_fill= PatternFill(start_color="690700", end_color="690700", fill_type="solid")
        white_font=Font(color="FFFFFF", bold=True)

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=4 + len(dias_laborales))):
            for cell in row:
                # Aplicar color verde a las celdas con "cobrado"
                if cell.value == 'cobrado':
                    cell.fill = green_fill
                # Aplicar color azul claro a las celdas del "Campus 2"
                elif cell.value == 'no-aplica':
                    cell.fill = gray_fill
                elif cell.value == 'no-cobrado':
                    cell.fill = red_fill
                elif cell.column_letter == 'A' and cell.value == 'Campus 2':
                    cell.fill = blue_fill
                
                # Columna de faltas totales
            faltas_cell = sheet.cell(row=row_idx + 2, column=4 + len(dias_laborales) + 1)
            grupo = sheet.cell(row=row_idx + 2, column=2).value  # Obtener grupo
            faltas_max = max_faltas.get(grupo, 0)
            if faltas_cell.value >= faltas_max:
                # Resaltar las celdas del becario y las faltas totales en rojo
                becario_cell=sheet.cell(row=row_idx + 2, column=3)
                becario_cell.fill = wine_fill  # Columna del becario
                becario_cell.font = white_font
                faltas_cell.fill = wine_fill
                faltas_cell.font = white_font

        writer.close()
        output.seek(0)

        # Guarda el archivo en la base de datos
        nombre_archivo = f'resumen_cobros_{ahora.strftime("%Y%m%d")}.xlsx'
        guardar_archivo_db(nombre_archivo, output.getvalue())

def obtener_resumen_data():
    with app.app_context():
        cursor = conexion.connection.cursor()
        
        cursor.execute("""
            SELECT b.campus, b.grupo, b.nombre, b.numcontrol, c.fecha
            FROM becarios b
            LEFT JOIN cobro_beca c ON b.numcontrol = c.numcontrol
            WHERE MONTH(c.fecha) = MONTH(CURDATE()) AND YEAR(c.fecha) = YEAR(CURDATE())
        """)
        
        resultados = cursor.fetchall()
        cursor.close()
        
        resumen = {}
        
        ahora = datetime.now()
        primer_dia = ahora.replace(day=1)
        ultimo_dia = ahora.replace(day=calendar.monthrange(ahora.year, ahora.month)[1])
        
        dias_laborales = set()
        while primer_dia <= ultimo_dia:
            if primer_dia.weekday() < 5:
                dias_laborales.add(primer_dia.day)
            primer_dia += timedelta(days=1)
        
        # Inicializa el resumen con todos los días del mes marcados como 'no-cobrado'
        for campus, grupo, nombre, numcontrol, fecha in resultados:
            if campus not in resumen:
                resumen[campus] = {}
            if grupo not in resumen[campus]:
                resumen[campus][grupo] = []
            
            dia_cobro = datetime.strptime(fecha, '%Y-%m-%d').day
            
            # Verifica si el becario ya está en la lista para el grupo
            becario = next((b for b in resumen[campus][grupo] if b['numcontrol'] == numcontrol), None)
            if not becario:
                becario = {
                    'nombre': nombre,
                    'numcontrol': numcontrol,
                    'dias_cobro': {dia: 'no-cobrado' for dia in range(1, ultimo_dia.day + 1)}
                }
                resumen[campus][grupo].append(becario)
            
            becario['dias_cobro'][dia_cobro] = 'cobrado'
        
        # Añadir días no aplicables y no cobrados
        dias_laborales_grupo = {
            'A': {0, 1, 2, 3, 4},  # Lunes a Viernes
            'B': {0, 2, 4},        # Lunes, Miércoles, Viernes
            'C': {1, 3}            # Martes, Jueves
        }
        for campus, grupos in resumen.items():
            for grupo, becarios in grupos.items():
                dias_cobro = dias_laborales_grupo.get(grupo, set())
                for becario in becarios:
                    for dia in range(1, ultimo_dia.day + 1):
                        fecha_actual = datetime(ahora.year, ahora.month, dia)
                        dia_semana = fecha_actual.weekday()

                        if fecha_actual > ahora:
                            estado = 'no-aplica'
                        elif dia_semana in dias_cobro:
                            estado = becario['dias_cobro'].get(dia, 'no-cobrado')
                            if estado == 'no-cobrado':
                                becario['dias_cobro'][dia] = 'no-cobrado'
                        else:
                            estado = 'no-aplica'
                        
                        if estado not in becario['dias_cobro']:
                            becario['dias_cobro'][dia] = estado
        

        return resumen

@app.route('/respaldos')
def listar_respaldos():
    cursor=conexion.connection.cursor()
    cursor.execute("SELECT id,nombre_archivo, fecha_creacion FROM respaldo_calendario ORDER BY fecha_creacion DESC")
    respaldos = cursor.fetchall()
    cursor.close()
    return render_template('respaldos.html', respaldos=respaldos)

@app.route('/eliminar_respaldos', methods=['POST'])
def eliminar_respaldos():
    respaldos_ids=request.form.getlist('respaldos')
    print(respaldos_ids)
    if respaldos_ids:
        cursor = conexion.connection.cursor()

        try:
            respaldos_ids= [int(id) for id in respaldos_ids]
            cursor.execute("DELETE FROM respaldo_calendario WHERE id IN (%s)" % ','.join(['%s'] * len(respaldos_ids)), tuple(respaldos_ids))
            conexion.connection.commit()
            flash('Respaldos eliminados correctamente.', 'success')
        except Exception as e:
            conexion.connection.rollback()
            flash(f'Error al eliminar respaldos: {e}', 'error')
        finally:
            cursor.close()
    
    return redirect(url_for('listar_respaldos'))

@app.route('/descargar_respaldo/<int:respaldo_id>')
def descargar_respaldo(respaldo_id):

    cursor = conexion.connection.cursor()
    cursor.execute("SELECT nombre_archivo, archivo FROM respaldo_calendario WHERE id = %s", (respaldo_id,))
    respaldo = cursor.fetchone()
    cursor.close()
    if respaldo:
        nombre_archivo, archivo = respaldo
        return send_file(
            io.BytesIO(archivo),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    else:
        return "Archivo no encontrado", 404
    
if __name__=='__main__':
    with app.app_context():
        iniciar_scheduler()
    app.run(debug=False)