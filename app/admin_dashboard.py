from flask import Blueprint, render_template, session, redirect, url_for, current_app, send_file
from datetime import datetime,date, timedelta
import calendar
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font



resumen_bp= Blueprint('resumen',__name__)


@resumen_bp.route('/resumen')
def resumen():
    conexion= current_app.config['MYSQL']
    ahora=datetime.now()
    primer_dia=ahora.replace(day=1)
    ultimo_dia=ahora.replace(day=calendar.monthrange(ahora.year, ahora.month)[1])
    mes_dias = list(range(1, ultimo_dia.day + 1))
    
    dias_laborales= []
    while primer_dia <= ultimo_dia:
        if primer_dia.weekday() < 5:
            dias_laborales.append(primer_dia.day)
        primer_dia += timedelta(days=1)
    

    try:
        cursor = conexion.connection.cursor()

        #Obtener becarios por grupo
        cursor.execute("SELECT nombre, numcontrol, grupo, campus FROM becarios WHERE activo = TRUE ORDER BY grupo, nombre, campus")

        becarios=cursor.fetchall()

        #Fechas de cobro
        cursor.execute("SELECT numcontrol,fecha FROM cobro_beca")
        cobros=cursor.fetchall()

        #Conversion Fecha
        cobros_dict={}
        for numcontrol,fecha in cobros:
            fecha_dt = datetime.strptime(fecha,'%Y-%m-%d').date()
            if numcontrol not in cobros_dict:
                cobros_dict[numcontrol]=set()
            cobros_dict[numcontrol].add(fecha_dt)

        #Estructura para resumen de cobro
        resumen_cobro = {}
        for nombre, numcontrol, grupo, campus in becarios:
            fechas_cobradas = cobros_dict.get(numcontrol, set())

            if not fechas_cobradas:
                continue
                        
            dias_cobro = set()
            if grupo == 'A':
                dias_cobro = {0, 1, 2, 3, 4}  # Lunes a Viernes
            elif grupo == 'B':
                dias_cobro = {0, 2, 4}  # Lunes, Miércoles, Viernes
            elif grupo == 'C':
                dias_cobro = {1, 3}  # Martes, Jueves
            
            faltas=0
            resume_cobro_dias = []
            for dia in dias_laborales:
                fecha_actual = date(ahora.year, ahora.month,dia)
                dia_semana=fecha_actual.weekday()
                
                if dia_semana in dias_cobro:
                    estado = 'cobrado' if fecha_actual in fechas_cobradas else 'no-cobrado' if fecha_actual <= ahora.date() else 'deshabilitado'
                    if estado == 'no-cobrado':
                        faltas += 1
                elif fecha_actual > ahora.date():
                    estado = 'deshabilitado'
                else:
                    estado = 'no-aplica'
                resume_cobro_dias.append({'dia':dia, 'estado':estado})
            
            if grupo == 'A' and faltas >=5 :
                faltas_totales= True
            elif grupo == 'B' and faltas >= 3:
                faltas_totales= True
            elif grupo == 'C' and faltas >= 2:
                faltas_totales= True
            else:
                faltas_totales= False
            if campus not in resumen_cobro:
                resumen_cobro[campus]={}
            if grupo not in resumen_cobro[campus]:
                resumen_cobro[campus][grupo]=[]
            resumen_cobro[campus][grupo].append({
                'nombre':nombre,
                'numcontrol':numcontrol,
                'dias_cobro':resume_cobro_dias,
                'faltas':faltas,
                'faltas_totales':faltas_totales
            })
            
        cursor.close()
        
        return render_template('admin_dashboard.html',resumen=resumen_cobro, mes_nombre=ahora.strftime('%B').title(), anio=ahora.year, mes_dias=dias_laborales)
    
    except Exception as e:
        return render_template('admin_dashboard.html', error=(e))
    
@resumen_bp.route('/descargar', methods=['POST'])
def descargar_lista():
    resumen_data=obtener_resumen_data()
    
    ahora= datetime.now()
    primer_dia=ahora.replace(day=1)
    ultimo_dia=ahora.replace(day=calendar.monthrange(ahora.year, ahora.month)[1])

    dias_laborales=[]
    while primer_dia<=ultimo_dia:
        if primer_dia.weekday()<5:
            dias_laborales.append(primer_dia.day)
        primer_dia += timedelta(days=1)

    rows=[]

    max_faltas={
        'A':5,
        'B':3,
        'C':2
    }
    for campus, grupos in resumen_data.items():
        for grupo, becarios in sorted(grupos.items(), key=lambda x: x[0]):
            becarios_ordenados = sorted(becarios, key=lambda b: b['nombre'])
            for becario in becarios_ordenados:
                fila = [campus, grupo, becario['nombre'], becario['numcontrol']]
                
                faltas=0
                # Añadir el estado de cada día del mes a la fila
                for dia in dias_laborales:  
                    estado_dia = 'no-aplica'
                    if dia in becario['dias_cobro']:
                        estado_dia = becario['dias_cobro'][dia]
                        if estado_dia == 'no-cobrado':
                            faltas += 1
                    fila.append(estado_dia)
                fila.append(faltas)
                rows.append(fila)

    # Crear nombres de columna dinámicamente para Excel
    column_names = ['Campus', 'Grupo', 'Becario', 'Número de Control'] + [dia for dia in dias_laborales] + ['Faltas']
    
    # Crear DataFrame con las filas y columnas correctas
    df = pd.DataFrame(rows, columns=column_names)
    output = io.BytesIO()
    writer=pd.ExcelWriter(output,engine='openpyxl')
    df.to_excel(writer,sheet_name='Resumen de Cobros',index=False)
    
    workbook = writer.book
    sheet = workbook['Resumen de Cobros']

    green_fill = PatternFill(start_color="00b900", end_color="00b900", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    gray_fill= PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
    red_fill= PatternFill(start_color="FF6F65", end_color="FF6F65", fill_type="solid")
    wine_fill= PatternFill(start_color="690700", end_color="690700", fill_type="solid")
    white_font=Font(color="FFFFFF", bold=True)

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=4 + len(dias_laborales))):
        for cell in row:
            # Aplicar color verde a las celdas con "cobrado"
            if cell.value == 'cobrado':
                cell.fill = green_fill
            # Aplicar color azul claro a las celdas del "Campus 2"
            elif cell.value == 'no-aplica':
                cell.fill = gray_fill
            elif cell.value == 'no-cobrado':
                cell.fill = red_fill
            elif cell.column_letter == 'A' and cell.value == 'Campus 2':
                cell.fill = blue_fill
            
            # Columna de faltas totales
        faltas_cell = sheet.cell(row=row_idx + 2, column=4 + len(dias_laborales) + 1)
        grupo = sheet.cell(row=row_idx + 2, column=2).value  # Obtener grupo
        faltas_max = max_faltas.get(grupo, 0)
        if faltas_cell.value >= faltas_max:
            # Resaltar las celdas del becario y las faltas totales en rojo
            becario_cell=sheet.cell(row=row_idx + 2, column=3)
            becario_cell.fill = wine_fill  # Columna del becario
            becario_cell.font = white_font
            faltas_cell.fill = wine_fill
            faltas_cell.font = white_font

    writer.close()
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'resumen_cobros_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

def obtener_resumen_data():
    
    conexion = current_app.config['MYSQL']
    cursor = conexion.connection.cursor()
    
    cursor.execute("""
        SELECT b.campus, b.grupo, b.nombre, b.numcontrol, c.fecha
        FROM becarios b
        LEFT JOIN cobro_beca c ON b.numcontrol = c.numcontrol
        WHERE MONTH(c.fecha) = MONTH(CURDATE()) AND YEAR(c.fecha) = YEAR(CURDATE())
    """)
    
    resultados = cursor.fetchall()
    cursor.close()
    
    resumen = {}
    
    ahora = datetime.now()
    primer_dia = ahora.replace(day=1)
    ultimo_dia = ahora.replace(day=calendar.monthrange(ahora.year, ahora.month)[1])
    
    dias_laborales = set()
    while primer_dia <= ultimo_dia:
        if primer_dia.weekday() < 5:
            dias_laborales.add(primer_dia.day)
        primer_dia += timedelta(days=1)
    
    # Inicializa el resumen con todos los días del mes marcados como 'no-cobrado'
    for campus, grupo, nombre, numcontrol, fecha in resultados:
        if campus not in resumen:
            resumen[campus] = {}
        if grupo not in resumen[campus]:
            resumen[campus][grupo] = []
        
        dia_cobro = datetime.strptime(fecha, '%Y-%m-%d').day
        
        # Verifica si el becario ya está en la lista para el grupo
        becario = next((b for b in resumen[campus][grupo] if b['numcontrol'] == numcontrol), None)
        if not becario:
            becario = {
                'nombre': nombre,
                'numcontrol': numcontrol,
                'dias_cobro': {dia: 'no-cobrado' for dia in range(1, ultimo_dia.day + 1)}
            }
            resumen[campus][grupo].append(becario)
        
        becario['dias_cobro'][dia_cobro] = 'cobrado'
    
    # Añadir días no aplicables y no cobrados
    dias_laborales_grupo = {
        'A': {0, 1, 2, 3, 4},  # Lunes a Viernes
        'B': {0, 2, 4},        # Lunes, Miércoles, Viernes
        'C': {1, 3}            # Martes, Jueves
    }
    for campus, grupos in resumen.items():
        for grupo, becarios in grupos.items():
            dias_cobro = dias_laborales_grupo.get(grupo, set())
            for becario in becarios:
                for dia in range(1, ultimo_dia.day + 1):
                    fecha_actual = datetime(ahora.year, ahora.month, dia)
                    dia_semana = fecha_actual.weekday()

                    if fecha_actual > ahora:
                        estado = 'no-aplica'
                    elif dia_semana in dias_cobro:
                        estado = becario['dias_cobro'].get(dia, 'no-cobrado')
                        if estado == 'no-cobrado':
                            becario['dias_cobro'][dia] = 'no-cobrado'
                    else:
                        estado = 'no-aplica'
                    
                    if estado not in becario['dias_cobro']:
                        becario['dias_cobro'][dia] = estado

    return resumen



